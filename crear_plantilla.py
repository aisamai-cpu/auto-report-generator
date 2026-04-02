import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ventas"

# Estilos
azul = PatternFill("solid", fgColor="2E86AB")
verde = PatternFill("solid", fgColor="27AE60")
gris = PatternFill("solid", fgColor="F0F4F8")
blanco = PatternFill("solid", fgColor="FFFFFF")
borde = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Titulo principal
ws.merge_cells('A1:E1')
ws['A1'] = '📊 Plantilla de Registro de Ventas'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = azul
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Instrucciones
ws.merge_cells('A2:E2')
ws['A2'] = 'Instrucciones: Agrega o elimina filas libremente. No modifiques los encabezados. La columna "mes" puede ser Enero, Febrero... o 2024, 2025, etc.'
ws['A2'].font = Font(italic=True, size=9, color='444444')
ws['A2'].fill = gris
ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 25

# Encabezados
encabezados = ['mes', 'producto', 'categoria', 'ventas', 'unidades']
for col, encabezado in enumerate(encabezados, 1):
    cell = ws.cell(row=3, column=col)
    cell.value = encabezado
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = verde
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = borde
ws.row_dimensions[3].height = 20

# Datos de ejemplo
datos = [
    ('Enero', 'Laptop', 'Tecnologia', 15000, 10),
    ('Enero', 'Mouse', 'Accesorios', 800, 40),
    ('Enero', 'Teclado', 'Accesorios', 1200, 30),
    ('Febrero', 'Laptop', 'Tecnologia', 18000, 12),
    ('Febrero', 'Monitor', 'Tecnologia', 9000, 9),
    ('Febrero', 'Mouse', 'Accesorios', 600, 30),
    ('Marzo', 'Laptop', 'Tecnologia', 21000, 14),
    ('Marzo', 'Teclado', 'Accesorios', 1600, 40),
    ('Marzo', 'Auriculares', 'Accesorios', 2400, 20),
]

for fila, dato in enumerate(datos, 4):
    fill = gris if fila % 2 == 0 else blanco
    for col, valor in enumerate(dato, 1):
        cell = ws.cell(row=fila, column=col)
        cell.value = valor
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borde
    ws.row_dimensions[fila].height = 18

# Ancho de columnas
anchos = [12, 18, 16, 12, 12]
for col, ancho in enumerate(anchos, 1):
    ws.column_dimensions[get_column_letter(col)].width = ancho

# Guardar
wb.save('data/plantilla_ventas.xlsx')
print("Plantilla creada: data/plantilla_ventas.xlsx")